"""Excel 报表生成服务"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Sequence

from app.models.issue import Issue
from app.models.test_case import TestCase


class ExcelService:
    """Excel 报表生成服务"""

    # 样式定义
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
    BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    SEVERITY_MAP = {"low": "低", "medium": "中", "high": "高", "critical": "严重"}
    ISSUE_STATUS_MAP = {"open": "待处理", "in_progress": "处理中", "resolved": "已解决", "closed": "已关闭"}
    PRIORITY_MAP = {"low": "低", "medium": "中", "high": "高"}
    CASE_STATUS_MAP = {"draft": "草稿", "active": "启用", "deprecated": "废弃"}

    @classmethod
    def _apply_header_style(cls, ws, row: int, col_count: int):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.HEADER_ALIGN
            cell.border = cls.BORDER

    @classmethod
    def _apply_cell_style(cls, ws, start_row: int, end_row: int, col_count: int):
        for row in range(start_row, end_row + 1):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = cls.CELL_ALIGN
                cell.border = cls.BORDER

    @classmethod
    def _auto_width(cls, ws, col_count: int, min_width: int = 12, max_width: int = 50):
        for col in range(1, col_count + 1):
            max_len = min_width
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)) * 2, max_width))
            ws.column_dimensions[get_column_letter(col)].width = max_len

    @classmethod
    def generate_issues_report(cls, issues: Sequence[Issue], projects: dict[int, str] | None = None) -> Workbook:
        """生成问题记录报表 - 按项目分组，每个项目单独一个Sheet

        Args:
            issues: 问题记录列表
            projects: {project_id: project_name} 映射，用于Sheet命名
        """
        if projects is None:
            projects = {}

        wb = Workbook()
        # 删除默认空Sheet
        default_ws = wb.active

        headers = ["ID", "标题", "描述", "严重程度", "状态", "报告人", "负责人", "创建时间", "更新时间"]

        # 按项目分组
        grouped: dict[int, list[Issue]] = {}
        no_project: list[Issue] = []
        for issue in issues:
            pid = issue.project_id
            if pid is not None:
                grouped.setdefault(pid, []).append(issue)
            else:
                no_project.append(issue)

        def _write_sheet(ws, issue_list: list[Issue]):
            """向Sheet写入问题数据"""
            ws.append(headers)
            cls._apply_header_style(ws, 1, len(headers))

            for issue in issue_list:
                ws.append([
                    issue.id, issue.title, issue.description,
                    cls.SEVERITY_MAP.get(issue.severity, issue.severity),
                    cls.ISSUE_STATUS_MAP.get(issue.status, issue.status),
                    issue.reporter, issue.assignee,
                    issue.created_at.strftime("%Y-%m-%d %H:%M") if issue.created_at else "",
                    issue.updated_at.strftime("%Y-%m-%d %H:%M") if issue.updated_at else "",
                ])

            if len(issue_list) > 0:
                cls._apply_cell_style(ws, 2, len(issue_list) + 1, len(headers))
            cls._auto_width(ws, len(headers))
            ws.freeze_panes = "A2"

        sheet_created = False

        # 按项目ID排序，每个项目创建一个Sheet
        for pid in sorted(grouped.keys()):
            issue_list = grouped[pid]
            project_name = projects.get(pid, f"项目{pid}")
            # Sheet名称最多31字符，替换非法字符
            safe_name = str(project_name).replace('/', '-').replace('\\', '-').replace('?', '-').replace('*', '-').replace('[', '(').replace(']', ')').replace(':', '-')
            if len(safe_name) > 28:
                safe_name = safe_name[:28]
            safe_name = f"{safe_name}({len(issue_list)}条)"

            if not sheet_created:
                # 第一个项目复用默认Sheet
                ws = default_ws
                ws.title = safe_name
                sheet_created = True
            else:
                ws = wb.create_sheet(title=safe_name)
            _write_sheet(ws, issue_list)

        # 未归类项目的问题
        if no_project:
            ws = wb.create_sheet(title="未归类项目")
            _write_sheet(ws, no_project)

        # 如果没有任何数据，保留一个空Sheet
        if not sheet_created and not no_project:
            default_ws.title = "问题记录"
            default_ws.append(headers)
            cls._apply_header_style(default_ws, 1, len(headers))
            cls._auto_width(default_ws, len(headers))
            default_ws.freeze_panes = "A2"

        return wb

    @classmethod
    def generate_test_cases_report(cls, cases: Sequence[TestCase]) -> Workbook:
        """生成测试用例报表"""
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        headers = ["ID", "项目ID", "标题", "描述", "前置条件", "步骤数", "优先级", "状态", "创建人", "创建时间"]
        ws.append(headers)
        cls._apply_header_style(ws, 1, len(headers))

        for case in cases:
            steps_text = ""
            if case.steps:
                for i, step in enumerate(case.steps, 1):
                    if isinstance(step, dict):
                        steps_text += f"{i}. {step.get('action', '')} (预期: {step.get('expected_result', '')})\n"
                    else:
                        steps_text += f"{i}. {step}\n"
            ws.append([
                case.id, case.project_id, case.title, case.description,
                case.preconditions, len(case.steps) if case.steps else 0,
                cls.PRIORITY_MAP.get(case.priority, case.priority),
                cls.CASE_STATUS_MAP.get(case.status, case.status),
                case.created_by_name,
                case.created_at.strftime("%Y-%m-%d %H:%M") if case.created_at else "",
            ])

        if len(cases) > 0:
            cls._apply_cell_style(ws, 2, len(cases) + 1, len(headers))
        cls._auto_width(ws, len(headers))

        ws.freeze_panes = "A2"
        return wb

    @classmethod
    def generate_summary_report(cls, issues: Sequence[Issue], cases: Sequence[TestCase], projects: dict[int, str] | None = None) -> Workbook:
        """生成汇总报表（问题按项目分Sheet + 测试用例Sheet + 统计汇总Sheet）"""
        wb = cls.generate_issues_report(issues, projects)

        # 新增测试用例 Sheet
        ws2 = wb.create_sheet(title="测试用例")
        headers = ["ID", "项目ID", "标题", "描述", "前置条件", "步骤数", "优先级", "状态", "创建人", "创建时间"]
        ws2.append(headers)
        cls._apply_header_style(ws2, 1, len(headers))

        for case in cases:
            ws2.append([
                case.id, case.project_id, case.title, case.description,
                case.preconditions, len(case.steps) if case.steps else 0,
                cls.PRIORITY_MAP.get(case.priority, case.priority),
                cls.CASE_STATUS_MAP.get(case.status, case.status),
                case.created_by_name,
                case.created_at.strftime("%Y-%m-%d %H:%M") if case.created_at else "",
            ])

        if len(cases) > 0:
            cls._apply_cell_style(ws2, 2, len(cases) + 1, len(headers))
        cls._auto_width(ws2, len(headers))
        ws2.freeze_panes = "A2"

        # 新增统计 Sheet
        ws3 = wb.create_sheet(title="统计汇总", index=0)
        ws3.append(["统计项", "数量"])
        cls._apply_header_style(ws3, 1, 2)
        ws3.append(["问题记录总数", len(issues)])
        ws3.append(["测试用例总数", len(cases)])

        # 问题状态统计
        for status_key, status_label in cls.ISSUE_STATUS_MAP.items():
            count = sum(1 for i in issues if i.status == status_key)
            ws3.append([f"问题-{status_label}", count])

        # 问题严重程度统计
        for sev_key, sev_label in cls.SEVERITY_MAP.items():
            count = sum(1 for i in issues if i.severity == sev_key)
            ws3.append([f"问题-{sev_label}", count])

        # 用例状态统计
        for status_key, status_label in cls.CASE_STATUS_MAP.items():
            count = sum(1 for c in cases if c.status == status_key)
            ws3.append([f"用例-{status_label}", count])

        cls._apply_cell_style(ws3, 2, ws3.max_row, 2)
        ws3.column_dimensions["A"].width = 25
        ws3.column_dimensions["B"].width = 15

        return wb
